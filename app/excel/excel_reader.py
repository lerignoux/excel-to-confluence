import logging
import re
from openpyxl import load_workbook

log = logging.getLogger("excel-to-confluence")


class ExcelReader():

    def __init__(self, config, bytes):
        self.config = config
        self.wb = load_workbook(bytes)

    def parse(self, sheet=None, header_row=None):
        if sheet is None:
            sheet = self.wb.get_sheet_names()[0]
        ws = self.wb.get_sheet_by_name(sheet)
        if header_row is None:
            header_row = self.best_header_row(ws)
        return {
            'sheets': self.get_sheets(),
            'data': self.parse_sheet(ws, header_row),
            'header_row': header_row,
            'sheet': sheet
        }

    def get_sheets(self):
        return self.wb.get_sheet_names()

    def best_header_row(self, ws):
        return int(ws.dimensions[1])

    def parse_sheet(self, ws, header_row):
        return {
            ws.title: [
                self.get_row(ws, row) for row in list(ws.iter_rows())[header_row-1:]
            ]
        }

    def get_row(self, ws, rows):
        """
        Extract the formatted sheet content
        """
        return [self.cell_json(cell) for cell in rows]

    def cell_json(self, cell):
        return {
            'value': cell.value,
            'style': {
                'color': self.excel_color(cell.font.color),
                'background-color': self.excel_color(cell.fill.bgColor),
                'font-size': cell.font.size,
                'font-weight': 'bold' if cell.font.bold else None,
                'font-style': 'italic' if cell.font.italic else None,
                'vertical-align': 'middle',
                'text-align': 'center'
            }
        }

    def excel_color(self, color):
        try:
            res = re.sub("^FF", "#", color.rgb)
            # log.info(res)
            return res
        except AttributeError:
            return "#000000"
