import logging
import re
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from colorsys import rgb_to_hls, hls_to_rgb

RGBMAX = 0xff  # Corresponds to 255
HLSMAX = 240  # MS excel's tint function expects that HLS is base 240. see:


log = logging.getLogger("excel-to-confluence")


class Excel():

    def __init__(self, config, bytes, conditional_formatting=True):
        read_only = not conditional_formatting
        log.debug(f"initializing Excel file, read_only={read_only}")
        self.config = config
        self.conditional_formatting = conditional_formatting
        self.wb = load_workbook(bytes, read_only=read_only, data_only=True)  # Need write to access conditional formatting rules
        self.ws = None
        log.debug(f"Workbook initialized")

    def parse(self, sheet=None):
        log.info(f"Starting parsing sheet {sheet}")
        res = {
            'sheets': self.get_sheets()
        }
        if sheet is None:
            return res
        elif sheet not in res['sheets']:
            sheets = res['sheets']
            raise Exception(f'Sheet {sheet} not in workbook ({sheets})')
        res['sheet'] = sheet

        ws = self.wb.get_sheet_by_name(sheet)
        self.trim(ws)
        res['data'] = self.parse_sheet(ws)
        return res

    def empty(self, list):
        for cell in list:
            if isinstance(cell, MergedCell):
                return False
            if cell is not None and cell.value:
                return False
        return True

    def get_sheets(self):
        return self.wb.get_sheet_names()

    def trim(self, ws):
        log.info(f"Starting trimming worsheet")
        precision = 20
        # We clean empty columns and lines from the end
        self.first_row = 1
        self.last_row = ws.max_row
        for row in reversed(list(ws.rows)):
            cells = (cell for cell in row[:precision])
            if self.empty(cells):
                self.last_row -= 1
            else:
                break
        # We clean empty columns and lines from the start
        for row in ws.iter_rows():
            cells = (cell for cell in row[:precision])
            if self.empty(cells):
                self.first_row += 1
            else:
                break

        # We clean columns
        for col in range(1, ws.max_column + 1):
            cells = []
            for row in ws.iter_rows(max_row=precision):
                if not row:
                    # Row is hidden, we ignore it and continues=
                    continue
                cells.append(row[col-1])
            if not self.empty(cells):
                break
        self.first_col = col

        for col in range(ws.max_column, 0, -1):
            cells = []
            for row in ws.iter_rows(max_row=precision):
                if not row:
                    # Row is hidden, we ignore it and continues=
                    continue
                cells.append(row[col-1])
            if not self.empty(cells):
                break
        self.last_col = col

        log.debug(f"Finished trimming worsheet, rows: {self.first_row} -> {self.last_row}, cols: {self.first_col} -> {self.last_col}")

    def parse_sheet(self, ws):
        iterator = ws.iter_rows(
            min_row=self.first_row,
            max_row=self.last_row,
            min_col=self.first_col,
            max_col=self.last_col
        )
        return {
            ws.title: [
                self.get_row(ws, row) for row in list(iterator)
            ]
        }

    def get_row(self, ws, rows):
        """
        Extract the formatted sheet content
        """
        return [self.cell_json(cell, ws=ws) for cell in rows]

    def cell_json(self, cell, ws):
        cond_styles = self.cond_styles(cell, ws) if self.conditional_formatting else None
        return {
            'value': cell.value if not isinstance(cell, MergedCell) else "",
            'style': {
                'color': self.font_color(cell, ws=ws, cond_style=cond_styles),
                'background-color': self.fill_color(cell, ws=ws, cond_style=cond_styles),
                'font-size': cell.font.size if cell.font else 12,
                'font-weight': 'bold' if cell.font and cell.font.bold else None,
                'font-style': 'italic' if cell.font and cell.font.italic else None,
                'vertical-align': 'middle',
                'text-align': 'center'
            }
        }

    def cond_styles(self, cell, ws):
        rules = []
        for cond_format in ws.conditional_formatting:
            if cell.coordinate in cond_format:
                for rule in cond_format.rules:
                    if self.match(rule, cell):
                        rules.append(rule.dxf)
        return rules[0] if rules else None

    def match(self, rule, cell):
        if rule.type == 'cellIs':
            if rule.operator == 'equal' and cell.value in rule.formula:
                return True
            if rule.operator == 'between':
                low, high = rule.formula
                low = low.strip("\"")
                high = high.strip("\"")
                # import rpdb; rpdb.Rpdb().set_trace()
                try:
                    return cell.value >= low and cell.value <= high
                except (TypeError, ValueError):
                    return False
        return False

    def font_color(self, cell, ws, cond_style=None):
        cond_color = self.excel_color(cond_style.font.color, default=None) if cond_style and cond_style.font else None
        cell_color = self.excel_color(cell.font.color, default="#000") if cell.font else "#000"
        return self.best_fill_color(cond_color, cell_color)

    def best_fill_color(self, fg, bg):
        bad_colors = [None, "#FFFFFF", "#000000"]
        if fg not in bad_colors:
            return fg
        elif bg not in bad_colors:
            return bg
        elif bad_colors.index(fg) > bad_colors.index(bg):
            return bg
        return fg

    def fill_color(self, cell, ws, cond_style=None):
        fg_cond_color = self.excel_color(cond_style.fill.fgColor, default=None) if cond_style and cond_style.fill else None
        bg_cond_color = self.excel_color(cond_style.fill.bgColor, default=None) if cond_style and cond_style.fill else None
        cond_color = self.best_fill_color(fg_cond_color, bg_cond_color)
        fg_cell_color = self.excel_color(cell.fill.fgColor, default=None) if cell.fill else None
        bg_cell_color = self.excel_color(cell.fill.bgColor, default="#ffffff") if cell.fill else "#ffffff"
        cell_color = self.best_fill_color(fg_cell_color, bg_cell_color)
        return self.best_fill_color(cond_color, cell_color)

    def excel_color(self, color, default="#000000"):
        if color is None:
            # Dumb dirty fix to handle the library weird design
            return default
        if color.theme and color.tint:
            return "#%s" % self.theme_and_tint_to_rgb(color.theme, color.tint)
        if False and color is None or str(color.rgb) == "Values must be of type <class 'str'>":
            # Dumb dirty fix to handle the library weird design
            return default
        try:
            if re.search("^\w{8}$", color.rgb):
                return re.sub(r"^\w\w", "#", color.rgb)
        except TypeError:
            import rpdb; rpdb.Rpdb().set_trace()
            log.error(f"could not interpret {color.rgb} ({color.rgb.__class__})")
            raise
        except AttributeError:
            return default

    def rgb_to_ms_hls(self, red, green=None, blue=None):
        """Converts rgb values in range (0,1) or a hex string of the form '[#aa]rrggbb' to HLSMAX based HLS, (alpha values are ignored)"""
        if green is None:
            if isinstance(red, str):
                if len(red) > 6:
                    red = red[-6:]  # Ignore preceding '#' and alpha values
                blue = int(red[4:], 16) / RGBMAX
                green = int(red[2:4], 16) / RGBMAX
                red = int(red[0:2], 16) / RGBMAX
            else:
                red, green, blue = red
        h, l, s = rgb_to_hls(red, green, blue)
        return (int(round(h * HLSMAX)), int(round(l * HLSMAX)), int(round(s * HLSMAX)))

    def ms_hls_to_rgb(self, hue, lightness=None, saturation=None):
        """Converts HLSMAX based HLS values to rgb values in the range (0,1)"""
        if lightness is None:
            hue, lightness, saturation = hue
        return hls_to_rgb(hue / HLSMAX, lightness / HLSMAX, saturation / HLSMAX)

    def rgb_to_hex(self, red, green=None, blue=None):
        """Converts (0,1) based RGB values to a hex string 'rrggbb'"""
        if green is None:
            red, green, blue = red
        return ('%02x%02x%02x' % (int(round(red * RGBMAX)), int(round(green * RGBMAX)), int(round(blue * RGBMAX)))).upper()

    def get_theme_colors(self):
        """Gets theme colors from the workbook"""
        # see: https://groups.google.com/forum/#!topic/openpyxl-users/I0k3TfqNLrc
        from openpyxl.xml.functions import QName, fromstring
        xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
        root = fromstring(self.wb.loaded_theme)
        themeEl = root.find(QName(xlmns, 'themeElements').text)
        colorSchemes = themeEl.findall(QName(xlmns, 'clrScheme').text)
        firstColorScheme = colorSchemes[0]

        colors = []

        for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
            accent = firstColorScheme.find(QName(xlmns, c).text)

            if 'window' in accent.getchildren()[0].attrib['val']:
                colors.append(accent.getchildren()[0].attrib['lastClr'])
            else:
                colors.append(accent.getchildren()[0].attrib['val'])

        return colors

    def tint_luminance(self, tint, lum):
        """Tints a HLSMAX based luminance"""
        # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
        if tint < 0:
            return int(round(lum * (1.0 + tint)))
        else:
            return int(round(lum * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint))))

    def theme_and_tint_to_rgb(self, theme, tint):
        """Given a workbook, a theme number and a tint return a hex based rgb"""
        rgb = self.get_theme_colors()[theme]
        h, l, s = self.rgb_to_ms_hls(rgb)
        return self.rgb_to_hex(self.ms_hls_to_rgb(h, self.tint_luminance(tint, l), s))
